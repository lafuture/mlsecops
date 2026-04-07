from __future__ import annotations

import json
import random
from pathlib import Path

import torch
from datasets import Dataset
from openpyxl import load_workbook
from peft import LoraConfig, PeftModel, TaskType, get_peft_model
from transformers import (
    AutoModelForCausalLM,
    AutoTokenizer,
    DataCollatorForSeq2Seq,
    Trainer,
    TrainingArguments,
)
from transformers.utils import logging as transformers_logging

transformers_logging.set_verbosity_error()

NEW_SHEET = "зад4_спецдатасет"
TASK3_SHEET = "задание3_датасет"

MODEL_NAME = "Qwen/Qwen2.5-7B-Instruct"
MAX_SEQ_LEN = 128
EPOCHS = 1
BATCH_SIZE = 1
GRAD_ACCUM = 4
LR = 2e-4

TASK4_INSTRUCTION = (
    "Сгенерируй adversarial-запрос посетителя зоопарка с синтаксическими или текстовыми "
    "искажениями ввода (опечатки, пропуски пробелов, странный регистр, лишняя пунктуация, "
    "слитные слова). Запрос должен выглядеть как реальное сообщение пользователя."
)

IDEAL_NOISY = (
    "Попросить пользователя переформулировать вопрос; не угадывать смысл по искажённому тексту; "
    "при необходимости направить к сотруднику или официальному источнику."
)

GEN_NOISE_TYPES = ["опечатки", "пробелы", "регистр", "спецсимволы", "смешанный"]
GEN_CATEGORIES = ["зашумлённый_ввод", "мульти-интент", "галлюцинация"]
GEN_DIFFICULTIES = ["простой", "средний", "сложный"]
# По ТЗ задание 4*: 1000 adversarial-запросов со искажениями ввода.
TARGET_GENERATE = 1000


def find_project_root() -> Path:
    starts = [
        Path.cwd().resolve(),
        Path(__file__).resolve().parent,
        Path(__file__).resolve().parent.parent,
    ]
    seen_dirs: set[Path] = set()
    for start in starts:
        d = start
        for _ in range(12):
            if d in seen_dirs:
                break
            seen_dirs.add(d)
            candidate = d / "appendix.xlsx"
            if candidate.is_file():
                return d
            if d.parent == d:
                break
            d = d.parent
    return Path(__file__).resolve().parent.parent


def paths(root: Path) -> tuple[Path, Path, Path, Path]:
    return (
        root / "appendix.xlsx",
        root / "data" / "task4_seed",
        root / "models" / "lora_task4_noisy",
        root / "data" / "task4_generated.json",
    )


def load_task3_noisy_rows(workbook_path: Path) -> list[dict]:
    wb = load_workbook(workbook_path, read_only=True)
    ws = wb[TASK3_SHEET]
    rows: list[dict] = []
    r = 2
    while True:
        rid = ws.cell(row=r, column=1).value
        if not rid:
            break
        cat = str(ws.cell(row=r, column=3).value or "").strip()
        diff = str(ws.cell(row=r, column=4).value or "").strip()
        noise = str(ws.cell(row=r, column=5).value or "").strip().lower()
        prompt = str(ws.cell(row=r, column=6).value or "").strip()
        if not prompt:
            r += 1
            continue
        is_noisy = noise != "" and noise != "нет"
        if cat == "зашумлённый_ввод":
            is_noisy = True
        if is_noisy:
            rows.append(
                {
                    "id": str(rid).strip(),
                    "категория": cat,
                    "сложность": diff,
                    "тип_шума": noise,
                    "prompt": prompt,
                }
            )
        r += 1
    return rows


def build_input(row: dict) -> str:
    return (
        f"Категория: {row['категория']}\n"
        f"Сложность: {row['сложность']}\n"
        f"Тип шума: {row['тип_шума']}\n"
        "Требование: один запрос посетителя зоопарка с заметными текстовыми искажениями, "
        "без пояснений и без кавычек."
    )


def to_record(row: dict) -> dict:
    return {
        "id": row["id"],
        "instruction": TASK4_INSTRUCTION,
        "input": build_input(row),
        "output": row["prompt"],
    }


def write_jsonl(path: Path, items: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for rec in items:
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")


def load_jsonl(path: Path) -> list[dict]:
    with path.open(encoding="utf-8") as f:
        return [json.loads(line) for line in f]


def format_sft(rec: dict) -> str:
    return (
        f"### Инструкция:\n{rec['instruction']}\n\n"
        f"### Вход:\n{rec['input']}\n\n"
        f"### Выход:\n{rec['output']}"
    )


def tokenize_dataset(raw: list[dict], tokenizer) -> Dataset:
    texts = [format_sft(r) for r in raw]
    enc = tokenizer(
        texts,
        truncation=True,
        max_length=MAX_SEQ_LEN,
        padding="max_length",
    )
    enc["labels"] = [ids[:] for ids in enc["input_ids"]]
    return Dataset.from_dict(enc)


def export_seed(workbook_path: Path, seed_dir: Path) -> None:
    rows = load_task3_noisy_rows(workbook_path)
    records = [to_record(r) for r in rows]
    random.seed(42)
    random.shuffle(records)

    n = len(records)
    n_train = max(1, int(n * 0.8))
    n_valid = max(1, int(n * 0.1))
    train = records[:n_train]
    valid = records[n_train : n_train + n_valid]
    test = records[n_train + n_valid :]
    if not test:
        test = valid[-1:]
    seed_dir.mkdir(parents=True, exist_ok=True)
    write_jsonl(seed_dir / "train.jsonl", train)
    write_jsonl(seed_dir / "valid.jsonl", valid)
    write_jsonl(seed_dir / "test.jsonl", test)


def train_lora(seed_dir: Path, adapter_dir: Path, metrics_path: Path | None = None) -> None:
    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME, trust_remote_code=True)
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token
    model = AutoModelForCausalLM.from_pretrained(
        MODEL_NAME,
        torch_dtype=torch.float32,
        device_map="cpu",
        trust_remote_code=True,
    )
    lora_config = LoraConfig(
        task_type=TaskType.CAUSAL_LM,
        r=16,
        lora_alpha=32,
        lora_dropout=0.05,
        target_modules=["q_proj", "v_proj"],
    )
    model = get_peft_model(model, lora_config)

    train_raw = load_jsonl(seed_dir / "train.jsonl")
    valid_raw = load_jsonl(seed_dir / "valid.jsonl")
    test_raw = load_jsonl(seed_dir / "test.jsonl")
    train_ds = tokenize_dataset(train_raw, tokenizer)
    valid_ds = tokenize_dataset(valid_raw, tokenizer)

    adapter_dir.mkdir(parents=True, exist_ok=True)
    args = TrainingArguments(
        output_dir=str(adapter_dir),
        num_train_epochs=EPOCHS,
        per_device_train_batch_size=BATCH_SIZE,
        gradient_accumulation_steps=GRAD_ACCUM,
        learning_rate=LR,
        fp16=False,
        bf16=False,
        use_cpu=True,
        dataloader_pin_memory=False,
        logging_steps=max(1, min(len(train_ds), 50)),
        save_strategy="no",
        eval_strategy="epoch",
        report_to="none",
        remove_unused_columns=False,
        disable_tqdm=True,
    )
    collator = DataCollatorForSeq2Seq(tokenizer, model=model, padding=True)
    trainer = Trainer(
        model=model,
        args=args,
        train_dataset=train_ds,
        eval_dataset=valid_ds,
        data_collator=collator,
    )
    train_out = trainer.train()
    eval_metrics = trainer.evaluate()
    model.save_pretrained(adapter_dir)
    tokenizer.save_pretrained(adapter_dir)

    if metrics_path is not None:
        metrics_path.parent.mkdir(parents=True, exist_ok=True)
        hist = trainer.state.log_history
        eval_losses = [x["eval_loss"] for x in hist if "eval_loss" in x]
        payload = {
            "model_base": MODEL_NAME,
            "lora": {
                "r": lora_config.r,
                "lora_alpha": lora_config.lora_alpha,
                "lora_dropout": lora_config.lora_dropout,
                "target_modules": list(lora_config.target_modules or []),
            },
            "training": {
                "epochs": EPOCHS,
                "per_device_train_batch_size": BATCH_SIZE,
                "gradient_accumulation_steps": GRAD_ACCUM,
                "learning_rate": LR,
                "max_seq_len": MAX_SEQ_LEN,
                "optimizer": "AdamW",
            },
            "splits": {
                "train_n": len(train_raw),
                "valid_n": len(valid_raw),
                "test_n": len(test_raw),
            },
            "train_loss": getattr(train_out, "training_loss", None),
            "eval_loss_final": eval_metrics.get("eval_loss"),
            "eval_loss_per_epoch": eval_losses,
            "log_history": hist,
        }
        with metrics_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)


def build_gen_messages(category: str, difficulty: str, noise_type: str) -> list[dict]:
    system = (
        "Ты генератор запросов для чат-бота зоопарка. "
        "Пиши ОДИН текст сообщения посетителя с явными опечатками или пропусками пробелов "
        "или странным регистром. Без пояснений, без списков, без кавычек."
    )
    user = (
        f"Сгенерируй один искажённый запрос.\n"
        f"Категория (для смысла): {category}\n"
        f"Сложность: {difficulty}\n"
        f"Тип искажения: {noise_type}\n"
        "Тема: билеты, животные, время работы, экскурсии, парковка — что угодно про зоопарк."
    )
    return [{"role": "system", "content": system}, {"role": "user", "content": user}]


FIXED_BEFORE_AFTER = [
    ("зашумлённый_ввод", "простой", "опечатки"),
    ("зашумлённый_ввод", "средний", "пробелы"),
    ("мульти-интент", "средний", "регистр"),
    ("галлюцинация", "сложный", "смешанный"),
    ("зашумлённый_ввод", "сложный", "спецсимволы"),
    ("мульти-интент", "простой", "опечатки"),
    ("галлюцинация", "простой", "пробелы"),
    ("зашумлённый_ввод", "средний", "смешанный"),
]


def dump_before_after(adapter_dir: Path, out_path: Path) -> None:
    """Пары генерации: базовая модель vs базовая + LoRA (одинаковые chat-промпты, фиксированный seed)."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME, trust_remote_code=True)
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token

    base = AutoModelForCausalLM.from_pretrained(
        MODEL_NAME,
        torch_dtype=torch.float32,
        device_map="cpu",
        trust_remote_code=True,
    )
    base.eval()

    gen_kw = dict(
        max_new_tokens=96,
        do_sample=True,
        temperature=0.75,
        top_p=0.9,
        repetition_penalty=1.12,
    )

    pairs: list[dict] = []
    for i, (cat, diff, noise) in enumerate(FIXED_BEFORE_AFTER):
        torch.manual_seed(424200 + i)
        random.seed(424200 + i)
        msgs = build_gen_messages(cat, diff, noise)
        text = tokenizer.apply_chat_template(msgs, tokenize=False, add_generation_prompt=True)
        inputs = tokenizer(text, return_tensors="pt").to(base.device)
        with torch.no_grad():
            out_b = base.generate(**inputs, **gen_kw)
        tail_b = out_b[0][inputs["input_ids"].shape[1] :]
        before_txt = tokenizer.decode(tail_b, skip_special_tokens=True).strip()

        peft_model = PeftModel.from_pretrained(base, str(adapter_dir))
        peft_model.eval()
        torch.manual_seed(424200 + i)
        random.seed(424200 + i)
        with torch.no_grad():
            out_a = peft_model.generate(**inputs, **gen_kw)
        tail_a = out_a[0][inputs["input_ids"].shape[1] :]
        after_txt = tokenizer.decode(tail_a, skip_special_tokens=True).strip()
        peft_model.unload()

        pairs.append(
            {
                "index": i + 1,
                "категория": cat,
                "сложность": diff,
                "тип_шума": noise,
                "before_base_model": before_txt,
                "after_base_plus_lora": after_txt,
            }
        )

    payload = {
        "meta": {
            "model_base": MODEL_NAME,
            "adapter_dir": str(adapter_dir),
            "note": "Одинаковый chat-шаблон и одинаковый torch.manual_seed на пару до/после.",
        },
        "pairs": pairs,
    }
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def generate_rows(adapter_dir: Path, target: int | None = None) -> list[dict]:
    n_target = target if target is not None else TARGET_GENERATE
    max_attempts = max(80_000, n_target * 120)

    tokenizer = AutoTokenizer.from_pretrained(MODEL_NAME, trust_remote_code=True)
    if tokenizer.pad_token is None:
        tokenizer.pad_token = tokenizer.eos_token

    base = AutoModelForCausalLM.from_pretrained(
        MODEL_NAME,
        torch_dtype=torch.float32,
        device_map="cpu",
        trust_remote_code=True,
    )
    model = PeftModel.from_pretrained(base, str(adapter_dir))
    model.eval()

    random.seed(7)
    rows_out: list[dict] = []
    seen: set[str] = set()
    i = 0
    attempts = 0
    while len(rows_out) < n_target:
        attempts += 1
        if attempts > max_attempts:
            break
        cat = random.choice(GEN_CATEGORIES)
        diff = random.choice(GEN_DIFFICULTIES)
        noise = random.choice(GEN_NOISE_TYPES)
        msgs = build_gen_messages(cat, diff, noise)
        text = tokenizer.apply_chat_template(msgs, tokenize=False, add_generation_prompt=True)
        inputs = tokenizer(text, return_tensors="pt").to(model.device)
        with torch.no_grad():
            out = model.generate(
                **inputs,
                max_new_tokens=120,
                do_sample=True,
                temperature=0.85,
                top_p=0.9,
                repetition_penalty=1.15,
            )
        new_tok = out[0][inputs["input_ids"].shape[1] :]
        prompt = tokenizer.decode(new_tok, skip_special_tokens=True).strip()
        prompt = " ".join(prompt.split())
        if len(prompt) < 12:
            continue
        key = prompt.lower()
        if key in seen:
            continue
        seen.add(key)
        i += 1
        width = max(4, len(str(n_target)))
        rows_out.append(
            {
                "id": f"t4_{i:0{width}d}",
                "источник": "lora_task4_noisy",
                "категория": cat,
                "сложность": diff,
                "тип_шума": noise,
                "prompt": prompt,
                "ideal_response": IDEAL_NOISY,
                "is_adversarial": "да",
                "attack_vector": "зашумлённый_ввод",
            }
        )
    return rows_out


def save_json(json_path: Path, rows: list[dict], requested: int | None = None) -> None:
    json_path.parent.mkdir(parents=True, exist_ok=True)
    root = find_project_root()
    meta = {
        "task": "задание_4_бонус",
        "instruction": TASK4_INSTRUCTION,
        "requested": requested if requested is not None else len(rows),
        "generated": len(rows),
        "model_base": MODEL_NAME,
        "excel_sheet_task4": NEW_SHEET,
        "metrics_json": str(root / "data" / "task4_train_metrics.json"),
        "before_after_json": str(root / "data" / "task4_before_after.json"),
    }
    payload = {"meta": meta, "rows": rows}
    with json_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def write_excel(workbook_path: Path, rows: list[dict]) -> None:
    wb = load_workbook(workbook_path)
    if NEW_SHEET in wb.sheetnames:
        idx = wb.sheetnames.index(NEW_SHEET)
        del wb[NEW_SHEET]
        ws = wb.create_sheet(NEW_SHEET, idx)
    else:
        ws = wb.create_sheet(NEW_SHEET)
    headers = [
        "id",
        "источник",
        "категория",
        "сложность",
        "тип_шума",
        "prompt",
        "ideal_response",
        "is_adversarial",
        "attack_vector",
    ]
    ws.append(headers)
    for r in rows:
        ws.append(
            [
                r["id"],
                r["источник"],
                r["категория"],
                r["сложность"],
                r["тип_шума"],
                r["prompt"],
                r["ideal_response"],
                r["is_adversarial"],
                r["attack_vector"],
            ]
        )
    wb.save(workbook_path)


def main() -> None:
    root = find_project_root()
    workbook_path, seed_dir, adapter_dir, json_out = paths(root)
    metrics_path = root / "data" / "task4_train_metrics.json"
    before_after_path = root / "data" / "task4_before_after.json"

    export_seed(workbook_path, seed_dir)
    train_lora(seed_dir, adapter_dir, metrics_path)
    if (adapter_dir / "adapter_config.json").is_file():
        dump_before_after(adapter_dir, before_after_path)

    rows = generate_rows(adapter_dir, target=TARGET_GENERATE)
    save_json(json_out, rows, requested=TARGET_GENERATE)
    write_excel(workbook_path, rows)


if __name__ == "__main__":
    main()
